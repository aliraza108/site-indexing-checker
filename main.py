import requests
import openpyxl
print(" 100 queries per day free")
def is_indexed_google(url: str, api_key: str, cx: str) -> bool:
    query = f"site:{url}"
    endpoint = "https://www.googleapis.com/customsearch/v1"
    params = {"key": api_key, "cx": cx, "q": query}
    r = requests.get(endpoint, params=params, timeout=10)
    r.raise_for_status()
    data = r.json()
    return "items" in data and len(data["items"]) > 0

def main():
    API_KEY = ""   # your API key
    CX = "90864d30dc0c54eeb"   # your CSE ID

    # Read sites from sites.txt
    with open("sites.txt", "r", encoding="utf-8") as f:
        sites = [line.strip() for line in f if line.strip()]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index Status"

    # Headers
    ws["A1"] = "Site"
    ws["B1"] = "Status"

    # Process each site
    for i, site in enumerate(sites, start=2):
        try:
            indexed = is_indexed_google(site, API_KEY, CX)
            status = "Indexed" if indexed else "Not Indexed"
        except Exception as e:
            status = f"Error: {e}"

        ws[f"A{i}"] = site
        ws[f"B{i}"] = status
        print(f"{site} -> {status}")

    # Save to Excel
    wb.save("output.xlsx")
    print("✅ Results saved to output.xlsx")

if __name__ == "__main__":
    main()
